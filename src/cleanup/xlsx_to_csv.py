import csv
import sys
from openpyxl import load_workbook


def xlsx_to_csv(xlsx_path: str, csv_path: str, sheet_name: str = None) -> None:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, lineterminator="\n")
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])

    wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(
            "Usage: uv run src/cleanup/xlsx_to_csv.py input.xlsx output.csv [sheet_name]"
        )
        sys.exit(1)

    sheet = sys.argv[3] if len(sys.argv) > 3 else None
    xlsx_to_csv(sys.argv[1], sys.argv[2], sheet)
