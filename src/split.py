# SPDX-FileCopyrightText: 2023 Sascha Brawer <sascha@brawer.ch>
# SPDX-License-Identifier: MIT

import argparse
import csv
import io
import os
import re
import urllib.request
import zipfile

import openpyxl
import PIL

from validator import COLUMNS, Validator

GIVEN_NAME_ABBREVS = {
    line.strip()
    for line in open(
        os.path.join(os.path.dirname(__file__), "given_name_abbreviations.csv")
    )
    if line != "Abbreviations\n"
}


# Company abbreviations, must be single words. If there's a need for
# multi-word sequences separated by whitespace, change the logic
# in split_company() below.
COMPANY_ABBREVS = {
    "AG",
    "A.-G.",
    "Cie.",
    "Co.",
    "Comp.",
    "Compagnie",
}


def is_valid_address(addr, validator):
    if m := re.match(r"^(.+) (\d+[a-t]?)$", addr):
        street, num = m.groups()
        return (street in validator.street_abbrevs) or (street in validator.streets)
    if addr in validator.pois:
        return True
    return False


def split(vol, validator):
    outpath = os.path.basename(vol).split(".")[0] + ".zip"
    zip_file = zipfile.ZipFile(outpath, "w")
    font = openpyxl.styles.Font(name="Calibri")
    red = openpyxl.styles.colors.Color(rgb="00FF2222")
    light_red = openpyxl.styles.colors.Color(rgb="00FFAAAA")
    red_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=red)
    light_red_fill = openpyxl.styles.fills.PatternFill(
        patternType="solid", fgColor=light_red
    )
    page_re = re.compile(r"^# Date: (\d{4}-\d{2}-\d{2}) Page: (\d+)/(.*)")
    sheet, date, page_id, lemma, name, row = None, None, None, "-", "", 0
    workbook, image = None, None
    for nline, line in enumerate(open(vol.replace("/proofread/", "/proofread/stage/"))):
        line = line.strip()
        input_pos = (os.path.basename(vol), nline + 1)
        if m := page_re.match(line):
            save_workbook(workbook, page_id, zip_file)
            date, page_id, page_num = m.groups()
            workbook = openpyxl.Workbook()
            del workbook["Sheet"]
            sheet = create_sheet(workbook, page_id, page_num)
            row = 2
            image = fetch_jpeg(page_id)
            continue
        p, pos, *score = line.split("#", 2)
        pos = ",".join([str(x) for x in simplify_pos(int(page_id), pos)])
        if score:
            # score = float(score[0].split('=')[1])
            name, givenname, occupation, address = [x.strip() for x in p.split(",")]
            if name == "-":
                name = lemma
            other = None
        else:
            p = [x.strip() for x in p.split(",")]
            p = [x for x in p if x != ""]
            if len(p) == 0:
                continue
            if p[0].startswith("-"):
                p[0] = lemma + " " + p[0][1:].strip()
            name, rest = split_family_name(p[0])
            # After "von Goumoens-von Tavel", the new lemma is "von Goumoens".
            lemma = name.split("-")[0].strip()
            company, rest = split_company(name, rest)
            if company:
                name = company
                maidenname = None
            else:
                maidenname, rest = split_maidenname(rest)
            p = [rest] + p[1:] if rest else p[1:]
            if company:
                title = "[Firma]"
            else:
                title, p = split_title(p, validator)
            address, address2, p = split_address(p, validator)
            givenname, p = split_givenname(p, validator)
            occupation, p = split_occupation(p, validator)
            other = ", ".join(p)
        row = row + 1

        entry = {
            "Scan": str(page_id),
            "ID": pos,
            "Name": name,
            "Vorname": givenname,
            "Ledigname": maidenname,
            "Adelsname": "",
            "Titel": title,
            "Beruf": occupation,
            "Beruf 2": "",
            "Adresse": address,
            "Adresse 2": address2,
            "Bemerkungen": other,
        }

        # TODO: This makes a lot of garbage pass without anyone noticing.
        # Disabling for now.
        #
        # if line has been preprocessed by scorer.py
        # it's considered to be good, don't revalidate
        #
        # if not score:
        #    bad = validator.validate(entry, input_pos)
        # else:
        #    bad = []
        bad = validator.validate(entry, input_pos)
        if other:
            # In the output of the splitting phase, but not elsewhere,
            # we consider the existence of remarks as "bad", leading
            # to human review of the entry.
            bad.add("Bemerkungen")

        for column_index, column in enumerate(COLUMNS):
            cell = sheet.cell(row, column_index + 1)
            cell.value = entry[column]
            cell.font = font
            if column in bad:
                cell.fill = red_fill
            elif len(bad) > 0:
                cell.fill = light_red_fill

        if True:  # set to False for speed-up in development
            cropped = crop_image(image, pos)
            sheet.add_image(cropped, f"B{row}")
            sheet.row_dimensions[row].height = cropped.height + 5

    save_workbook(workbook, page_id, zip_file)
    zip_file.close()


def save_workbook(workbook, page_id, zip_file):
    if workbook is None:
        return
    workbook.save("tmp.xlsx")
    with open("tmp.xlsx", "rb") as f:
        content = f.read()
    os.remove("tmp.xlsx")
    zip_file.writestr(
        f"{page_id}.xlsx",
        content,
        compress_type=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    )


def split_company(name, rest):
    words = rest.split()
    for i, word in enumerate(words):
        if word in COMPANY_ABBREVS:
            company = " ".join([name] + words[: i + 1])
            return company, " ".join(words[i + 1 :])
    return None, rest


def split_family_name(n):
    n = n.replace(" -", "-").replace("- ", "-")
    words = n.split()
    pos = 0
    prefixes = {"de", "De", "von", "Von", "v.", "V."}
    if words[0] in prefixes:
        pos = pos + 1
    if any(words[pos].endswith("-" + p) for p in prefixes):
        pos = pos + 1
    return (" ".join(words[: pos + 1]), " ".join(words[pos + 1 :]))


def split_givenname(p, validator):
    if len(p) == 0:
        return ("", [])
    if all(n in validator.given_names for n in p[0].split()):
        return (p[0], p[1:])
    else:
        return ("", p)


def split_maidenname(n):
    if n.startswith("geb.") or n.startswith("gb."):
        words = n.split()
        if len(words) >= 2:
            if words[1] in {"v.", "V.", "von", "Von"} and len(words) >= 3:
                return ("von " + words[2], " ".join(words[3:]))
            else:
                return (words[1], " ".join(words[2:]))
    return ("", n)


def split_title(p, validator):
    if len(p) > 0 and p[0] in validator.titles:
        return (p[0], p[1:])
    else:
        return ("", p)


def split_address(p, validator):
    if len(p) == 0:
        return ("", "", [])
    if is_valid_address(p[0], validator):
        return (p[0], "", p[1:])
    if is_valid_address(p[-1], validator):
        return (p[-1], "", p[0:-1])
    last = p[-1].removesuffix(".")
    if m := re.match(r"(.+\d+) ([a-t])", last):
        addr, rest = ("".join(m.groups()), p[:-1])
    elif last and last[-1] in "0123456789":
        addr, rest = (last, p[:-1])
    else:
        return ("", "", p)
    if m := re.match(r"^(.+) (\d+\s?[a-t]?) u\. (\d+\s?[a-t]?)$", addr):
        street, housenum1, housenum2 = m.groups()
        return (f"{street} {housenum1}", f"{street} {housenum2}", rest)
    elif m := re.match(r"^(.+ \d+\s?[a-z]?) u\. (.+)$", addr):
        a1, a2 = m.groups()
        return (a1, a2, rest)
    return (addr, "", rest)


def split_occupation(p, validator):
    if len(p) == 0:
        return ("", p)
    occ = p[0]
    occ = {"Nent.": "Rent."}.get(occ, occ)
    if occ in validator.occupations:
        return (occ, p[1:])
    # Sometimes OCR (or the typesetter) missed a final dot, as in "Schneid"
    if not occ.endswith(".") and occ + "." in validator.occupations:
        return (occ + ".", p[1:])
    return ("", p)


# Returns the file paths to all address book volumes.
def list_volumes():
    path = os.path.join(os.path.dirname(__file__), "..", "proofread")
    path = os.path.normpath(path)
    return sorted(
        [os.path.join(path, f) for f in os.listdir(path) if f.endswith(".txt")]
    )


def create_sheet(workbook, page_id, page_num):
    sheet = workbook.create_sheet(page_num)
    font = openpyxl.styles.Font(name="Calibri", bold=True)
    gray = openpyxl.styles.colors.Color(rgb="00DDDDDD")
    gray_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=gray)
    cell = sheet.cell(1, 1)
    cell.value = f"https://www.e-rara.ch/bes_1/periodical/pageview/{page_id}"
    cell.font = font
    cell.fill = gray_fill
    sheet.merge_cells("A1:J1")
    for i, col in enumerate(COLUMNS):
        cell = sheet.cell(2, i + 1)
        cell.value = col
        cell.font = font
        cell.fill = gray_fill
    sheet.column_dimensions["A"].width = 3  # ID
    sheet.column_dimensions["B"].width = 35  # Scan
    return sheet


# {page_id: [(x, y, width, height), ...], ...}
page_columns = None


def simplify_pos(page_id, pos):
    global page_columns
    if page_columns is None:
        page_columns = read_page_columns()
    columns = page_columns.get(page_id, [])
    if not columns:
        columns = [(200, 150, 900, 2500), (980, 150, 900, 2500)]

    left, top, right, bottom = 2000, 2977, 0, 0
    for p in pos.split(";"):
        x, y, w, h = [int(x) for x in p.split(",")]
        if h < 0:
            h = 100
        left, right = min(x, left), max(x + w, right)
        top, bottom = min(y, top), max(y + h, bottom)
    left, right = (300, 1050) if left < 600 else (1000, 1750)
    top, bottom = max(top - 5, 0), min(bottom + 5, 2977)

    # Align to column boundaries detected with Computer Vision,
    # but only if the column does not span the entire page width.
    if right - left < 1200:
        mid_x = left + (right - left) // 2
        mid_y = top + (bottom - top) // 2
        for x, y, w, h in columns:
            if x <= mid_x <= x + w and y <= mid_y <= y + h:
                if x + w + 30 < 2000:
                    w += 30
                return (x, top, w, bottom - top)

    # Fallback for text boxes spanning both columns.
    return (left, top, right - left, bottom - top)


# Load the content of data/page_columns.csv into memory.
def read_page_columns():
    result = {}
    data_path = os.path.join(os.path.dirname(__file__), "..", "data")
    with open(os.path.join(data_path, "page_columns.csv"), "r") as fp:
        for row in csv.DictReader(fp):
            page_id = int(row["PageID"])
            x, y = int(row["X"]), int(row["Y"])
            w, h = int(row["Width"]), int(row["Height"])
            result.setdefault(page_id, []).append((x, y, w, h))
    return result


# Fetch the JPEG image for a single page from e-rara.ch.
# If the image is already in cache, the cached content is returned
# without re-downloading it over the internet.
def fetch_jpeg(pageid):
    filepath = os.path.join("cache", "images", f"{pageid}.jpg")
    if not os.path.exists(filepath):
        os.makedirs(os.path.join("cache", "images"), exist_ok=True)
        url = f"https://www.e-rara.ch/download/webcache/2000/{pageid}"
        print(f"fetching {url}")
        with urllib.request.urlopen(url) as u:
            img = u.read()
        # Write to a temp file, followed by (atomic) rename, so we never
        # have partially written files in the final location, even if
        # the process crashes while writing out the file.
        with open(filepath + ".tmp", "wb") as f:
            f.write(img)
        os.rename(filepath + ".tmp", filepath)
    return PIL.Image.open(filepath)


def crop_image(img, pos):
    x, y, w, h = [int(val) for val in pos.split(",")]
    crop = img.reduce(3, (x, y, x + w, y + h))
    out = io.BytesIO()
    crop.save(out, format="PNG")
    return openpyxl.drawing.image.Image(out)


# Parse the string passed as the --years command-line argument.
# For example, "1880-1883,1905" --> {1880, 1881, 1882, 1883, 1905}.
def parse_years(years):
    result = set()
    for s in years.split(","):
        s = s.strip()
        if "-" in s:
            if m := re.match(r"(\d{4})\-(\d{4})", s):
                for year in range(int(m.group(1)), int(m.group(2)) + 1):
                    result.add(year)
            else:
                raise ValueError(f"not in YYYY-YYYY format: {s}")
        else:
            result.add(int(s))
    return result


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--years", default="1860", type=parse_years)
    args = ap.parse_args()
    validator = Validator()
    for vol in list_volumes():
        year = int(os.path.basename(vol)[:4])
        if year in args.years:
            split(vol, validator)
    validator.report()
