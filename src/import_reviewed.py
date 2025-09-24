# SPDX-FileCopyrightText: 2024 Sascha Brawer <sascha@brawer.ch>
# SPDX-License-Identifier: MIT

# Imports one manually reviewed address book volume into the `reviewed` folder.
#
# For example, running this command:
#
#   venv/bin/python3 import_reviewed.py ~/Downloads/1860-02-01_reviewed.zip
#
# will write the file `1860-02-01.csv` into the `reviewed` subdirectory.
# This tool does not perform any checks against lists of known good names,
# but build_release.py runs this kind of checks when generating release data.

import argparse
import csv
import openpyxl
import os
import re
import zipfile

from validator import COLUMNS, Validator


# Known typos in column names; we fix them here instead of editing the data.
COLUMN_TYPOS = {
    "Addresse": "Adresse",
    "Addresse 2": "Adresse 2",
    "Bemerkung": "Bemerkungen",
    "Bermerkungen": "Bemerkungen",
}


IGNORED_COLUMNS = {
    "Scan",  # scan image in review Excels; scan ID in our output
    "Unklar",  # input for human reviewers before reviewing
}


def process_zip(path):
    volume_pattern = re.compile(r"^(\d{4}-\d{2}-\d{2})_reviewed\.zip$")
    if m := volume_pattern.match(path.rsplit("/", 1)[-1]):
        volume = m.group(1)
    else:
        assert "filename should be like YYYY-MM-DD_reviewed.zip", path
    outpath = os.path.join("reviewed", "%s.csv" % volume)
    files = []
    page_pattern = re.compile(
        r"^https://www.e-rara.ch/bes_1/periodical/pageview/(\d+)$"
    )
    pattern = re.compile(r"^.*/([0-9]+)_reviewed.*\.xlsx$")
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            if m := pattern.match(info.filename):
                files.append((int(m.group(1)), info.filename))
        if len(files) == 0:
            print('No files in ZIP file match "*/{pageid}_reviewed.xlsx"')
            return
    with zipfile.ZipFile(path) as zf, open(outpath, "w") as outfile:
        writer = csv.writer(outfile)
        writer.writerow(COLUMNS)
        fatal_errors = False
        for _, filename in sorted(files):
            wb = openpyxl.load_workbook(
                data_only=True,
                filename=zf.open(filename, mode="r"),
                keep_links=False,
                keep_vba=False,
                read_only=True,
            )
            sheets = [s for s in wb]
            assert len(sheets) == 1, filename
            sheet = sheets[0]
            page_match = page_pattern.match(sheet.cell(1, 1).value)
            if not page_match:
                assert "cannot find page_id in cell A1", filename
            page_id = page_match.group(1)
            col = {}
            for i in range(1, sheet.max_column + 1):
                if title := sheet.cell(row=2, column=i).value:
                    title = COLUMN_TYPOS.get(title, title)
                    if title not in IGNORED_COLUMNS:
                        if title not in COLUMNS:
                            print(f'{filename}: unknown column "{title}"')
                            fatal_errors = True
                            continue
                        col[title] = i
            for row in sheet.iter_rows(min_row=3, values_only=True):
                entry = {}
                for col_title, col_index in col.items():
                    if value := row[col_index - 1]:
                        entry[col_title] = str(value).strip()
                entry["Scan"] = page_id
                outrow = [""] * len(COLUMNS)
                for i, col_title in enumerate(COLUMNS):
                    if value := entry.get(col_title):
                        outrow[i] = value
                writer.writerow(outrow)
    if fatal_errors:
        return
    validator = Validator()
    with open(outpath) as fp:
        line = 0
        for row in csv.DictReader(fp):
            line += 1
            validator.validate(row, pos=(outpath, line))
    validator.report()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("zipfile")
    args = parser.parse_args()
    process_zip(args.zipfile)
