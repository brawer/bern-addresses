# SPDX-FileCopyrightText: 2024 Sascha Brawer <sascha@brawer.ch>
# SPDX-License-Identifier: MIT
#
# Tool for checking a single Excel sheet from human reviewers.

import argparse
import re
import sys

import openpyxl

import validator

IGNORED_COLUMNS = {
    "Scan",  # scan image in review Excels; scan ID in our output
    "Unklar",  # input for human reviewers before reviewing
}


def read_reviewed_excel(filename):
    wb = openpyxl.load_workbook(
        data_only=True,
        filename=filename,
        keep_links=False,
        keep_vba=False,
        read_only=True,
    )
    sheets = [s for s in wb]
    if len(sheets) != 1:
        print(f'file "{filename}" should have only 1 sheet')
        sys.exit(1)
    sheet = sheets[0]
    page_pattern = re.compile(
        r"^https://www.e-rara.ch/bes_1/periodical/pageview/(\d+)$"
    )
    page_match = page_pattern.match(sheet.cell(1, 1).value)
    if not page_match:
        print("cannot find page_id in cell A1")
        sys.exit(1)
    page_id = page_match.group(1)
    col = {}
    bad = False
    for i in range(1, sheet.max_column + 1):
        if title := sheet.cell(row=2, column=i).value:
            if title in IGNORED_COLUMNS:
                continue
            if title not in validator.COLUMNS:
                print(f'unknown column: "{title}"')
                bad = True
                continue
            col[title] = i
    if bad:
        sys.exit(1)
    for row in sheet.iter_rows(min_row=3, values_only=True):
        entry = {}
        for col_title, col_index in col.items():
            value = row[col_index - 1] or ""
            entry[col_title] = value
        entry["Scan"] = page_id
        yield entry


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("file")
    args = parser.parse_args()
    if not args.file.endswith(".xlsx"):
        print("Usage: python3 src/check_reviewed.py excel_sheet.xlsx")
        sys.exit(1)
    val = validator.Validator()
    line = 3
    for entry in read_reviewed_excel(args.file):
        val.validate(entry, (args.file, line))
        line += 1
    val.report()
