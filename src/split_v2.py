# SPDX-FileCopyrightText: 2025 Sascha Brawer <sascha@brawer.ch>
# SPDX-License-Identifier: MIT
#
# Tool for splitting OCR-ed lines. Generates Excel files for human review.
# For quick debugging, pass --format=text.
#
# Usage:
#
#     venv/bin/python3 src/split_v2.py --years=1870-1878 --format=excel
#     venv/bin/python3 src/split_v2.py --years=1865 --format=text


from argparse import ArgumentParser
import csv
from dataclasses import dataclass
from layout import LayoutAnalysis
import io
import os
from pathlib import Path
import re
import zipfile

import cv2 as cv
import numpy as np
import openpyxl

from utils import (
    AddressBookEntry,
    Box,
    OCRLine,
    Page,
    parse_years,
    read_pages,
    read_ocr_lines,
)
from validator import COLUMNS, Validator


REPLACEMENTS = {
    "..": ".,",
    "'": "’",
    "ẵ": "ä",
    "å": "ä",
    "Bolwerk": "Bollwerk",
    "Casinoplag": "Casinoplatz",
    "gaffe": "gasse",
    "I.": "J.",
    "Igfr.": "Jgfr.",
    "Inkg.": "Jnkg.",
    "Junferng": "Junkerng",
    "Käshdir": "Käshdlr",
    "Megg": "Metzg",
    "Mezg": "Metzg",
    "Nabbenth": "Rabbenth",
    "Nent": "Rent",
    "Nevifor": "Revisor",
    "plazg": "platzg",
    "plaz ": "platz ",
    "plagg": "platzg",
    "Räfich": "Käfich",
    "Regt.": "Negt.",
    "Ressler": "Kessler",
    "SchauFlagg": "Schauplatzg",
    "Schlsfr": "Schlssr",
    "Schweft": "Schwest",
}


COMPANY_ABBREVS = {
    "AG",
    "A.-G.",
    "Cie.",
    "Co.",
    "Comp.",
    "Compagnie",
    "Gebr.",
    "Gebrüder",
    "Gebrüd.",
    "& Cie.",
    "& Co.",
    "& Comp.",
    "& Cp.",
    "& Sohn",
    "& Söhne",
    "u. Cie.",
    "u. Comp.",
}


MAIDEN_NAME_PREFIXES = {
    "geb.",
    "gb.",
    "geborne",
    "geborene",
}


NOBILITY_PREFIXES = {
    "de": "de",
    "De": "de",
    "von": "von",
    "Von": "von",
    "v.": "von",
    "V.": "von",
}


class Splitter:
    def __init__(self, validator: Validator):
        self.validator = validator

    def split(self, lines: list[OCRLine]) -> list[AddressBookEntry]:
        result = []
        columns = set((l.page_id, l.column) for l in lines)
        for page_id, col in sorted(columns):
            col_lines = [l for l in lines if l.page_id == page_id and l.column == col]
            result.extend(self.split_column(col_lines))
        return result

    def split_column(self, lines: list[OCRLine]) -> list[AddressBookEntry]:
        result: list[AddressBoookEntry] = []
        lemma: str = ""
        lines = merge_lines(sorted(lines, key=lambda l: l.box.y))
        min_x = min(line.box.x for line in lines)
        max_x = max(line.box.x + line.box.width for line in lines)
        for line in lines:
            name, rest = self.split_name(line.text)
            if name == "—":
                name = lemma
            else:
                # After "von Goumoens-von Tavel", the new lemma is "von Goumoens".
                lemma = name.split("-")[0].strip()
            company, rest = self.split_company(name, rest)
            if company:
                name = company
                maiden_name = ""
                title = "[Firma]"
            else:
                maiden_name, rest = self.split_maiden_name(rest)
                title, rest = self.split_title(rest)
            given_name, rest = self.split_given_name(rest)
            if not title:  # title sometimes before, sometimes after given name
                title, rest = self.split_title(rest)
            addresses, rest = self.split_addresses(rest)
            occupations, rest = self.split_occupations(rest)
            box = Box(
                x=min_x, y=line.box.y, width=max_x - min_x, height=line.box.height
            )
            entry = AddressBookEntry(
                id=None,
                page_id=line.page_id,
                box=box,
                family_name=name,
                given_name=given_name,
                maiden_name=maiden_name,
                nobility_name="",
                title=title,
                occupations=occupations,
                addresses=addresses,
                workplace="",
                unrecognized=rest,
            )
            result.append(entry)
        return result

    def split_name(self, text: str) -> (str, str):
        p = text.split(",")
        n = p[0].replace(" -", "-").replace("- ", "-")
        words = n.split()
        pos = 0
        if words[0] in NOBILITY_PREFIXES:
            pos = pos + 1
        if any(words[pos].endswith("-" + p) for p in NOBILITY_PREFIXES):
            pos = pos + 1
        words[0] = NOBILITY_PREFIXES.get(words[0], words[0])
        name = " ".join(words[: pos + 1])
        if name in "-–—":
            name = "—"
        rest_list = [" ".join(words[pos + 1 :])] + p[1:]
        rest = ", ".join(rp.strip() for rp in rest_list if rp)
        return (name, rest)

    def split_company(self, name: str, rest: str) -> (str, str):
        words = rest.split()
        for a in COMPANY_ABBREVS:
            if rest.startswith(a):
                company = f"{name} {a}"
                rest = rest.removeprefix(a).removeprefix(",").strip()
                return company, rest
        return "", rest

    def split_maiden_name(self, text: str) -> (str, str):
        if not any(text.startswith(p) for p in MAIDEN_NAME_PREFIXES):
            return "", text
        parts = text.split(",")
        words = parts[0].split()[1:]
        maiden_name, rest_words = words[0], words[1:]
        if len(words) >= 2:
            if nob := NOBILITY_PREFIXES.get(words[0]):
                maiden_name, rest_words = nob + " " + words[1], words[2:]
        p = [" ".join(rest_words)] + parts[1:] if rest_words else parts[1:]
        rest = ", ".join([x.strip() for x in p])
        return (maiden_name, rest)

    def split_title(self, text: str) -> (str, str):
        for title in self.validator.titles:
            if text.startswith(title):
                rest = text.removeprefix(title).strip().removeprefix(",").strip()
                return (title, rest)
        return ("", text)

    def split_given_name(self, text: str) -> (str, str):
        parts = [p.strip() for p in text.split(",")]
        if len(parts) > 0:
            if all(n in self.validator.given_names for n in parts[0].split()):
                return parts[0], ", ".join(parts[1:])
        return "", text

    def split_occupations(self, text: str) -> (list[str], str):
        all_occupations = self.validator.occupations
        parts = [p.strip() for p in text.split(",")]
        found, rest = [], []
        for p in parts:
            if p in all_occupations:
                found.append(p)
                continue
            if p + "." in all_occupations:
                # Sometimes OCR (or the typesetter) missed a final dot,
                # as in "Schneid".
                found.append(p + ".")
                continue
            if m := re.match(r"(.+) (u\.|und) (.+)", p):
                p1, _und, p2 = m.groups()
                if p1 in all_occupations and p2 in all_occupations:
                    found.append(p1)
                    found.append(p2)
                    continue
            rest.append(p)
        return found, ", ".join(rest)

    def split_addresses(self, text: str) -> (list[str], str):
        p = [p.strip() for p in text.split(",")]
        if len(p) == 0:
            return [], text
        addrs = self.cleanup_address(p[0])
        if len(addrs) > 0:
            p = p[1:]
        if len(p) > 0:
            final_addrs = self.cleanup_address(p[-1])
            if final_addrs:
                p = p[:-1]
                addrs.extend(final_addrs)
        return addrs, ", ".join(p)

    def cleanup_address(self, addr: str) -> list[str]:
        val = self.validator
        is_street = lambda s: (s in val.street_abbrevs) or (s in val.streets)
        if addr in val.pois:
            return [addr]
        if m := re.match(r"^(.+) (\d+[a-t]?)\.?$", addr):
            street, num = m.groups()
            if is_street(street):
                return [f"{street} {num}"]
        if m := re.match(r"^(.+) (\d+[a-t]?)\s?(u\.|und) (\d+[a-t]?)$", addr):
            street, num1, _und, num2 = m.groups()
            if is_street(street):
                return [f"{street} {num1}", f"{street} {num2}"]
        if m := re.match(r"^(.+) (\d+[a-t]?)\s?(u\.|und) (.+) (\d+[a-t]?)$", addr):
            s1, num1, _und, s2, num2 = m.groups()
            if is_street(s1) and is_street(s2):
                return [f"{s1} {num1}", f"{s2} {num2}"]
        if m := re.match(r"^(.+) (\d+[a-t]?)\s?(u\.|und)\s?(.+)$", addr):
            s1, num1, _und, poi = m.groups()
            if is_street(s1) and poi in val.pois:
                return [f"{s1} {num1}", poi]
        return []


def main(years: set[int], format: str) -> None:
    validator = Validator()
    splitter = Splitter(validator)
    for volume, volume_pages in sorted(read_pages().items()):
        year = int(volume[:4])
        if year not in years:
            continue
        if format == "excel":
            out_zip = zipfile.ZipFile(f"{volume}.zip", "w")
        else:
            out_zip = None
        ocr_lines = read_ocr_lines(volume)
        for page in volume_pages:
            lines = [l for l in ocr_lines if l.page_id == page.id]
            split_lines = splitter.split(lines)
            match format:
                case "text":
                    for entry in split_lines:
                        if entry.unrecognized:
                            print(entry.unrecognized)
                case "excel":
                    workbook = make_excel_workbook(page, split_lines, validator)
                    workbook.save("tmp.xlsx")
                    with open("tmp.xlsx", "rb") as f:
                        content = f.read()
                    os.remove("tmp.xlsx")
                    out_zip.writestr(
                        f"{page.id}.xlsx",
                        content,
                        compress_type=zipfile.ZIP_DEFLATED,
                        compresslevel=9,
                    )
        if out_zip is not None:
            out_zip.close()


def make_excel_workbook(
    page: Page, entries: list[AddressBookEntry], validator: Validator
) -> openpyxl.Workbook:
    columns = [col for col in COLUMNS if col != "ID"]
    la = LayoutAnalysis(page)
    page_image = la.rotated_image

    font = openpyxl.styles.Font(name="Calibri")
    bold_font = openpyxl.styles.Font(name="Calibri", bold=True)
    red = openpyxl.styles.colors.Color(rgb="00FF2222")
    light_red = openpyxl.styles.colors.Color(rgb="00FFAAAA")
    red_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=red)
    light_red_fill = openpyxl.styles.fills.PatternFill(
        patternType="solid", fgColor=light_red
    )
    gray = openpyxl.styles.colors.Color(rgb="00DDDDDD")
    gray_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=gray)

    workbook = openpyxl.Workbook()
    del workbook["Sheet"]

    # Add first row.
    sheet = workbook.create_sheet(page.label)
    row = 1
    cell = sheet.cell(1, row)
    cell.value = f"https://www.e-rara.ch/bes_1/periodical/pageview/{page.id}"
    cell.font = font
    cell.fill = gray_fill
    sheet.merge_cells("A1:M1")

    # Add second row.
    row += 1
    for i, col in enumerate(columns):
        cell = sheet.cell(row, i + 1)
        cell.value = col
        cell.font = bold_font
        cell.fill = gray_fill
    sheet.column_dimensions["A"].width = 3  # ID
    sheet.column_dimensions["B"].width = 35  # Scan
    sheet.column_dimensions["O"].width = 20  # nicht zuweisbar

    # Add remaining rows.
    for entry in entries:
        row += 1
        entry_dict = entry.to_dict()
        bad = validator.validate(entry_dict, pos=None)
        if entry_dict.get("nicht zuweisbar"):
            bad.add("nicht zuweisbar")
        for column_index, column in enumerate(columns):
            cell = sheet.cell(row, column_index + 1)
            cell.value = entry_dict[column]
            cell.font = font
            if column in bad:
                cell.fill = red_fill
            elif len(bad) > 0:
                cell.fill = light_red_fill
        cropped = crop_image(page_image, entry.box)
        sheet.add_image(cropped, f"B{row}")
        sheet.row_dimensions[row].height = cropped.height

    return workbook


def crop_image(image: np.ndarray, box: Box) -> openpyxl.drawing.image.Image:
    cropped = image[box.y : box.y + box.height, box.x : box.x + box.width]
    size = (int(box.width / 2.5), int(box.height / 2.5))
    scaled = cv.resize(cropped, size, interpolation=cv.INTER_CUBIC)
    ok, png = cv.imencode(".png", scaled)
    if not ok:
        raise ValueError(f"could not encode PNG image at {box}")
    out = io.BytesIO()
    out.write(png.tobytes())
    return openpyxl.drawing.image.Image(out)


def merge_lines(lines: list[OCRLine]) -> list[OCRLine]:
    result: list[OCRLine] = []
    if len(lines) == 0:
        return result

    assert all(l.page_id == lines[0].page_id for l in lines)
    assert all(l.column == lines[0].column for l in lines)
    column_x = min(l.box.x for l in lines)
    last = None
    for line in lines:
        last = result[-1] if len(result) > 0 else None
        x = line.box.x - column_x
        text = cleanup_text(line.text)
        if x > 200 and re.match(r"^[A-Z]\.$", line.text):
            continue
        if x < 70 or last is None:
            result.append(
                OCRLine(
                    page_id=line.page_id, column=line.column, text=text, box=line.box
                )
            )
            continue
        last_text = last.text
        sep = " " if any(text.startswith(p) for p in ("und", "u.")) else ""
        if any(last_text.endswith(c) for c in "⸗-="):
            text = last_text[:-1] + sep + text
        else:
            text = " ".join((last_text + " " + text).split())
        box = last.box.union(line.box)
        merged = OCRLine(page_id=line.page_id, column=line.column, text=text, box=box)
        result[-1] = merged
    return result


def cleanup_text(s: str) -> str:
    if s[-1] in ":=":
        s = s[:-1] + "-"
    s = s.replace("ſ", "s").replace("ß", "ss").replace("⸗", "-")
    s = re.sub(r"\.([0-9A-ZÄÖÜ])", lambda m: ". " + m.group(1), s)
    s = re.sub(r"([0-9] [a-z][,.]?)", lambda m: m.group(1).replace(" ", ""), s)

    for a, b in REPLACEMENTS.items():
        s = s.replace(a, b)
    return s


if __name__ == "__main__":
    ap = ArgumentParser()
    ap.add_argument("--years", default="1864", type=parse_years)
    ap.add_argument("--format", default="excel", choices=["excel", "text"])
    args = ap.parse_args()
    main(years=args.years, format=args.format)
